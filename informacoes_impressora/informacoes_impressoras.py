#Importando todas as classes necessarias
from pysnmp.hlapi.v3arch.asyncio import * # importa a classe para chamar o SNMP
from openpyxl.styles import Font, Border, Side, Alignment # importa as classes para formatar a celula
from openpyxl.utils import get_column_letter # importa uma função para converter um numero inteiro em coluna do excel
from openpyxl import Workbook # impora a clase para manupular a tabela
import asyncio # importa a biblioteca para escrever codigo assincrono 
import sys # importa a biblioteca sys para testes

def pegar_informacoes(ips):
    impressoras = []
    erros_comunicacao = []
    erros_status = []

    async def main():
        for ip in ips:
            # Cria uma função para pegar as informções via CMD utilizando os OID do protocolo SNMP
            resultado = await get_cmd(
                SnmpEngine(),
                CommunityData("public"),

                await UdpTransportTarget.create((ip, 161)),
                ContextData(),

                ObjectType(ObjectIdentity("1.3.6.1.2.1.1.1.0")),          # Descrição
                ObjectType(ObjectIdentity("1.3.6.1.2.1.1.5.0")),          # Nome do Dispostivo
                ObjectType(ObjectIdentity("1.3.6.1.2.1.1.6.0")),          # Local
                ObjectType(ObjectIdentity("1.3.6.1.2.1.43.5.1.1.17.1")),  # Serial
                ObjectType(ObjectIdentity("1.3.6.1.2.1.43.10.2.1.4.1.1")) # Contador
            )

            error_indication, error_status, error_index, var_binds = resultado

            # Pega todos os IPs que deram erro de comunicação
            if error_indication:
                erros_comunicacao.append(ip)
                continue

            # Pega todos os IPs que deram erro de status
            if error_status:
                erros_status.append(ip)
                continue

            # Cria o dicionario com as informações recebidas
            dados = {
                "ip": ip,
                "modelo": str(var_binds[0][1]).split(";", 1)[0],
                "departamento": str(var_binds[1][1]),
                "local": str(var_binds[2][1]),
                "serial": str(var_binds[3][1]),
                "contador": int(var_binds[4][1]),
            }

            # Adcionona  o dicionario na listas de impressoras
            impressoras.append(dados)

    # Roda a função main de forma assincrona para a coleta dos dados
    asyncio.run(main())

    # Retorna com as informações de impressoras e seus erros
    return impressoras, erros_comunicacao, erros_status

# lista de IPs de todas as impressoras 
ips = [
    "192.168.8.52",
    "192.168.8.70",
    "192.168.8.86",
    "192.168.8.90",
    "192.168.8.109",
    "192.168.8.246",
    "192.168.8.186",
    "192.168.8.187",
    "192.168.8.134",
    "192.168.8.136",
    "192.168.8.240",
    "192.168.8.141",
    "192.168.8.230",
    "192.168.8.231",
    "192.168.8.188",
    "192.168.8.247",
    "192.168.8.244",
    "192.168.8.252",
    "192.168.8.251",
    "192.168.8.119",
    "192.168.8.145",
    "192.168.8.243",
    "192.168.8.248",
    "192.168.8.253",
    "192.168.7.79",
    "192.168.7.100",
    "192.168.7.101",
    "192.168.7.103",
    "192.168.7.197",
    "192.168.70.30",
    "192.168.70.31",
    "192.168.40.31",
    "192.168.40.30",
    "192.168.40.32"
]

# Criação da tabela do excel 
wb = Workbook()
ws = wb.active

# Chamando a função para pegar todas as informações e a aloca par uma variavel
dados = pegar_informacoes(ips)

# Separa as informações em 3 novas variaveis
impressoras = dados[0]
erros_comunicacao = dados[1]
erros_status = dados[2]

# Pega a posição das colunas
colunas = {
    "B": "local",
    "C": "departamento",
    "D": "modelo",
    "E": "ip",
    "F": "serial",
    "G": "contador",
}

# Troca o titulo da primeira aba
ws.title = "Impressoras"

# Definindo o titulos da tabela
ws['B2'] = "Local"
ws['C2'] = "Departamento"
ws['D2'] = "Impressora"
ws['E2'] = "IP"
ws['F2'] = "N. de Série"
ws['G2'] = "C. Inicial"
ws['H2'] = "C. Final"

# Variavel para definir a linha inicial
linha = 3

# Adicionando informações da impressoras nas celulas
for impressora in impressoras:
    for coluna, chave in colunas.items():
        ws[f"{coluna}{linha}"] = impressora.get(chave, "")

    linha += 1

# Criando uma nova aba 
ws2 = wb.create_sheet("Erros de Comunicação")

# Difinindo o titulo
ws2['B2'] = "IP com Falha"

# Resetando valor da linha
linha = 3

# Criando uma listagem com os IPs com erros
for erro in erros_comunicacao:
    ws2[f"B{linha}"] = erro

    linha += 1

#------------------------Formatação------------------------

# Pegando a quantidade de linhas
qtd_lws = len(ips) - len(erros_comunicacao) + 2
qtd_lws2 = len(erros_comunicacao) + 2

# Criando um estilo para bordas
borda_fina = Border(
    left = Side(style="thin"),
    right = Side(style="thin"),
    top = Side(style="thin"),
    bottom = Side(style="thin")
)

# Criando um estilo para o alinhamento
alinhar = Alignment(
    horizontal = "center",
    vertical = "center"
)

# Criando um estilo para a fonte
fonte = Font(
    name = "calibri",
    size = "12"
)

for linha in ws.iter_rows(min_row=2, max_row=qtd_lws, min_col=2, max_col=8):
    for celula in linha:
        celula.border = borda_fina
        celula.font = fonte
        celula.alignment = alinhar

# Difinindo a linha 2 como fonte bold
for celula in ws[2]:
    celula.font = Font(bold = True)

for linha in ws2.iter_rows(min_row=2, max_row=qtd_lws2, min_col=2, max_col=2):
    for celula in linha:
        celula.border = borda_fina
        celula.font = fonte
        celula.alignment = alinhar

ws2["B2"].font = Font(bold = True)

# Salvando a tabela com o nome
wb.save("relatorio_contador_impressoras.xlsx")